from __future__ import print_function
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook, Workbook
import concurrent.futures
import time
import multiprocessing
import threading

def get_document_type(mime_type):
    if mime_type == 'application/vnd.google-apps.document':
        return 'Google Docs'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
        return 'Microsoft Word'
    elif mime_type == 'application/rtf':
        return 'Rich Text Format'
    elif mime_type == 'text/html':
        return 'HTML'
    elif mime_type == 'application/pdf':
        return 'PDF'
    elif mime_type == 'application/x-zip-compressed':
        return "ZIP"
    elif mime_type == 'application/vnd.google-apps.folder':
        return 'Folder'
    elif mime_type == 'image/png':
        return 'PNG'
    elif mime_type == 'text/xml':
        return 'XML'
    elif mime_type == 'application/java-archive':
        return 'JAR'
    elif mime_type == 'application/vnd.google-apps.shortcut':
        return 'SHORTCUT'
    elif mime_type.startswith("video/"):
        return "VIDEO"
    elif mime_type == 'application/epub+zip':
        return 'EPUB'
    elif mime_type == 'application/zip':
        return 'ZIP'
    elif mime_type == 'application/vnd.google-apps.spreadsheet':
        return 'Google Sheets'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        return 'Microsoft Excel'
    elif mime_type == 'application/vnd.google-apps.presentation':
        return 'Google Slides'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.presentationml.presentation':
        return 'Microsoft PowerPoint'
    elif mime_type == 'text/plain':
        return 'Plain text'
    elif mime_type == 'application/vnd.oasis.opendocument.text':
        return 'OpenDocument Text'
    elif mime_type == 'application/vnd.oasis.opendocument.spreadsheet':
        return 'OpenDocument Spreadsheet'
    elif mime_type == 'application/vnd.oasis.opendocument.presentation':
        return 'OpenDocument Presentation'
    elif mime_type.startswith("image/"):
        return "Image"
    else:
        return mime_type


def search_file(email):
    """Search file in drive location"""
    key_path = "maerifat.json"

    creds = None
    if os.path.exists(key_path):
        creds = service_account.Credentials.from_service_account_file(
            key_path, scopes=["https://www.googleapis.com/auth/drive"],
            subject=email
        )

    try:
        service = build('drive', 'v3', credentials=creds)
        folders = []
        files_data = []

        if os.path.exists("april.xlsx"):
            wb = load_workbook(filename="april.xlsx")
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "File Data"
            ws.append(["File Name", "File Type", "Size", "Owners", "Shared With", "View Link", "Created Time", "Modified Time", "Last Viewed Time"])

        page_token = None
        while True:

            response = service.files().list(q="'me' in owners",
                                spaces='drive',
                                fields='nextPageToken, files(id, name, mimeType, size, owners, webViewLink, permissions, createdTime, modifiedTime, viewedByMeTime)',
                                pageToken=page_token).execute()
            time.sleep(2)
            for file in response.get('files', []):

                owners = file.get('owners', [])
                owner_emails = [owner.get('emailAddress') for owner in owners]
                shared_with = []
                permissions = file.get('permissions', [])
               
                print(file.get("name"))
                
                for permission in permissions:
                    print("here we go")
                    print(permission)
                    if ('emailAddress' in permission) and (len(permission['emailAddress']) > 2) :
                        if not permission['emailAddress'].endswith(("@domain1.com","@domain2.in","@domain3.co")):
                            shared_with.append(permission['emailAddress'].strip())
                    if 'anyone' in permission['type']:
                        print(file.get("name"))
                        print("has excessive permissions") 
                        shared_with.append("anyonewithLink".strip())
                            
                # print(F'Found file: {file.get("name")}, {file.get("id")}, owned by: {", ".join(owner_emails)}, shared with: {",".join(shared_with)}, view link: {file.get("webViewLink")}')

                if shared_with:
                    try:
                        size = int(str(file.get("size")))

                        if size < 1024:
                            size_str = "{} bytes".format(size)
                        elif size < 1024**2:
                            size_str = "{:.2f} KB".format(size/1024)
                        elif size < 1024**3:
                            size_str = "{:.2f} MB".format(size/1024**2)
                        else:
                            size_str = "{:.2f} GB".format(size/1024**3)
                    except :
                        size_str = "Unknown"

                    mime_type= file.get("mimeType")

                    files_data.append((file.get("name"), get_document_type(mime_type), size_str, ", ".join(owner_emails), " , ".join(shared_with).strip(), file.get("webViewLink"), file.get("createdTime"), file.get("modifiedTime"), file.get("viewedByMeTime")))

            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

        for data in files_data:
            ws.append(data)

        wb.save(f"April.xlsx")
    
    except HttpError as error:
        print(F'An error occurred: {error}')
        return None

# with open('users.txt', 'r') as file:
#   emails = [email.strip() for email in file.readlines()]
#   print(emails)

with open('users.txt', 'r') as file:
    emails = [email.strip() for email in file]
    print(emails)

start_time = time.time()

for email in emails:
    try:
        search_file(email)
        print("good")
    except:
        print("bad")

end_time = time.time()

execution_time = start_time - end_time
print(execution_time)
